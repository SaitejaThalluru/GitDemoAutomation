import openpyxl

book = openpyxl.load_workbook("C:\\Users\\ThalluruSaiteja\\Downloads\\Exceldemo.xlsx")
sheet = book.active
# Reading a value
cell = sheet.cell(row=2, column=1)
print(cell.value)
# writing a value in excel
sheet.cell(row=3, column=4).value = "sai"
print(sheet.cell(row=3, column=4).value)
# Total number of rows in a sheet
print(sheet.max_row)
# Total number of columns in a sheet
print(sheet.max_column)
# shortcut to print value
print(sheet['A1'].value)
# print all values in excel file
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

